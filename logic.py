# logic.py
import os
import json
import openpyxl
import pandas as pd
from schema import (
    OUTPUT_PATH, PROGRESS_FILE, get_config, DATA_START_ROW
)

def initialize_system(q_count: int):
    """確保雲端環境下有 Excel 檔案與正確的表頭"""
    if not os.path.exists(OUTPUT_PATH):
        create_new_workbook(q_count)
        # 清除舊的進度記錄
        if os.path.exists(PROGRESS_FILE):
            os.remove(PROGRESS_FILE)
    else:
        # 檢查現有檔案的問題數量是否符合
        wb = openpyxl.load_workbook(OUTPUT_PATH)
        ws = wb.active
        # 簡單檢查：找最後一個 Q{n} 是否正確
        config = get_config(q_count)
        last_q_col = config["question_starts"][-1] + 1
        expected_header = f"Q{q_count}"
        if ws.cell(row=3, column=last_q_col).value != expected_header:
            return False # 表示不匹配
    return True

def create_new_workbook(q_count: int):
    """動態產生一個全新的 Excel 活頁簿，並自動設定好表頭"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "問卷資料"
    
    ws.cell(row=1, column=1, value="問卷資料收集結果")
    
    config = get_config(q_count)
    
    for q_idx, start_col in enumerate(config["question_starts"]):
        base_col = start_col + 1
        ws.cell(row=3, column=base_col, value=f"Q{q_idx+1}")
        option_labels = ["非常滿意(1)", "滿意(2)", "尚可(3)", "不滿意(4)", "非常不滿意(5)"]
        for offset, label in enumerate(option_labels):
            ws.cell(row=4, column=base_col + offset, value=label)
            
    ws.cell(row=4, column=config["text_col_1"] + 1, value="反映與建議")
    ws.cell(row=4, column=config["text_col_2"] + 1, value="其他建議")
    
    wb.save(OUTPUT_PATH)
    return wb

def load_progress() -> int:
    """讀取進度 JSON 檔"""
    if os.path.exists(PROGRESS_FILE) and os.path.exists(OUTPUT_PATH):
        try:
            with open(PROGRESS_FILE, "r", encoding="utf-8") as f:
                return json.load(f).get("next_row", DATA_START_ROW)
        except:
            pass
    return DATA_START_ROW

def save_progress(next_row: int):
    """將目前的進度列數存入 JSON"""
    with open(PROGRESS_FILE, "w", encoding="utf-8") as f:
        json.dump({"next_row": next_row}, f)

def reset_data():
    """清除全部資料與進度檔"""
    if os.path.exists(OUTPUT_PATH):
        os.remove(OUTPUT_PATH)
    if os.path.exists(PROGRESS_FILE):
        os.remove(PROGRESS_FILE)

def get_workbook():
    """取得現有的 Excel 物件"""
    return openpyxl.load_workbook(OUTPUT_PATH)

def get_excel_download_bytes() -> bytes:
    """以二進位讀取 Excel，供前端下載使用"""
    if os.path.exists(OUTPUT_PATH):
        with open(OUTPUT_PATH, "rb") as f:
            return f.read()
    return b""

def save_survey_entry(wb, current_row: int, choices: list, t1: str, t2: str, q_count: int):
    """將單筆問卷資料寫入 Excel"""
    ws = wb.active
    config = get_config(q_count)
    
    # 寫入選擇題
    for q_idx, val in enumerate(choices):
        start_col = config["question_starts"][q_idx] + 1
        for offset in range(5):
            ws.cell(row=current_row, column=start_col + offset).value = (
                1 if (offset + 1) == val else None
            )
            
    # 寫入文字題
    ws.cell(row=current_row, column=config["text_col_1"] + 1).value = t1 or None
    ws.cell(row=current_row, column=config["text_col_2"] + 1).value = t2 or None
    
    # 儲存 Excel 與寫入下一筆的進度
    wb.save(OUTPUT_PATH)
    save_progress(current_row + 1)

def update_and_get_stats(wb, current_row: int, q_count: int) -> pd.DataFrame | None:
    """計算百分比、更新 Excel 分頁並回傳 Pandas DataFrame 給 UI 顯示"""
    ws_data = wb.active
    config = get_config(q_count)
    
    if "統計結果" in wb.sheetnames:
        ws_stats = wb["統計結果"]
    else:
        ws_stats = wb.create_sheet("統計結果")

    total_responses = current_row - DATA_START_ROW
    if total_responses <= 0:
        # 如果沒有資料，清除舊的統計分頁內容（如果有）
        if "統計結果" in wb.sheetnames:
            for row in ws_stats.iter_rows():
                for cell in row:
                    cell.value = None
        return None

    headers =["問題", "非常滿意(1)", "滿意(2)", "尚可(3)", "不滿意(4)", "非常不滿意(5)"]
    for col_num, header in enumerate(headers, 1):
        ws_stats.cell(row=1, column=col_num).value = header

    stats_data =[]

    for q_idx, start_col_idx in enumerate(config["question_starts"]):
        q_label = f"Q{q_idx+1}"
        row_stats = [q_label]
        
        for opt_offset in range(5):
            count = 0
            for r in range(DATA_START_ROW, current_row):
                if ws_data.cell(row=r, column=start_col_idx + opt_offset + 1).value == 1:
                    count += 1
            
            percentage = (count / total_responses) * 100
            row_stats.append(f"{percentage:.1f}%")

        for col_idx, val in enumerate(row_stats, 1):
            ws_stats.cell(row=q_idx + 2, column=col_idx).value = val
            
        stats_data.append(row_stats)

    wb.save(OUTPUT_PATH)
    return pd.DataFrame(stats_data, columns=headers)
