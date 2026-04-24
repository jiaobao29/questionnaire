import streamlit as st
import openpyxl
import os
import json
import pandas as pd

# ── CONFIGURATION ────────────────────────────────────────────────────────────
OUTPUT_PATH   = "questionnaire_filled.xlsx"
PROGRESS_FILE = "progress.json"

QUESTIONS_COUNT = 4
OPTIONS_RANGE = "12345"

# Column Mapping (0-indexed base logic from original)
QUESTION_COL_START =[1, 6, 11, 16] 
TEXT_COL_1 = 36   # Excel col 37 (AK)
TEXT_COL_2 = 37   # Excel col 38 (AL)
DATA_START_ROW = 5 

# ── PAGE CONFIGURATION ───────────────────────────────────────────────────────
st.set_page_config(page_title="問卷快速輸入系統", page_icon="🚀", layout="wide")

# ── FILE & PROGRESS MANAGEMENT ───────────────────────────────────────────────
def create_new_workbook():
    """動態產生一個全新的 Excel 活頁簿，並自動設定好表頭"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "問卷資料"
    
    # 建立基礎表頭，讓格式符合原本從第 5 列開始寫入 (DATA_START_ROW = 5) 的邏輯
    ws.cell(row=1, column=1, value="問卷資料收集結果")
    
    # 設定選擇題表頭
    for q_idx, start_col in enumerate(QUESTION_COL_START):
        base_col = start_col + 1
        ws.cell(row=3, column=base_col, value=f"Q{q_idx+1}")
        for offset in range(5):
            ws.cell(row=4, column=base_col + offset, value=f"選項{offset+1}")
            
    # 設定文字題表頭
    ws.cell(row=4, column=TEXT_COL_1 + 1, value="反映與建議")
    ws.cell(row=4, column=TEXT_COL_2 + 1, value="其他建議")
    
    wb.save(OUTPUT_PATH)
    return wb

def load_progress():
    # 只有當 Excel 存在時才讀取進度，否則一律從頭開始
    if os.path.exists(PROGRESS_FILE) and os.path.exists(OUTPUT_PATH):
        try:
            with open(PROGRESS_FILE, "r", encoding="utf-8") as f:
                return json.load(f).get("next_row", DATA_START_ROW)
        except:
            pass
    return DATA_START_ROW

def save_progress(row):
    with open(PROGRESS_FILE, "w", encoding="utf-8") as f:
        json.dump({"next_row": row}, f)

# ── STATISTICS ENGINE ────────────────────────────────────────────────────────
def update_and_get_stats(wb, current_row):
    """計算百分比、更新 Excel 分頁並回傳 Pandas DataFrame 給網頁顯示"""
    ws_data = wb.active
    
    if "統計結果" in wb.sheetnames:
        ws_stats = wb["統計結果"]
    else:
        ws_stats = wb.create_sheet("統計結果")

    total_responses = current_row - DATA_START_ROW
    if total_responses == 0:
        return None

    headers =["問題", "非常滿意(1)", "滿意(2)", "尚可(3)", "不滿意(4)", "非常不滿意(5)"]
    for col_num, header in enumerate(headers, 1):
        ws_stats.cell(row=1, column=col_num).value = header

    stats_data =[]

    # 計算每一題
    for q_idx, start_col_idx in enumerate(QUESTION_COL_START):
        q_label = f"Q{q_idx+1}"
        row_stats = [q_label]
        
        for opt_offset in range(5):
            count = 0
            for r in range(DATA_START_ROW, current_row):
                if ws_data.cell(row=r, column=start_col_idx + opt_offset + 1).value == 1:
                    count += 1
            
            percentage = (count / total_responses) * 100
            row_stats.append(f"{percentage:.1f}%")

        # 寫入 Excel
        for col_idx, val in enumerate(row_stats, 1):
            ws_stats.cell(row=q_idx + 2, column=col_idx).value = val
            
        stats_data.append(row_stats)

    wb.save(OUTPUT_PATH)
    
    # 轉換為 DataFrame 給 Streamlit 顯示
    df = pd.DataFrame(stats_data, columns=headers)
    return df

# ── MAIN ENGINE ──────────────────────────────────────────────────────────────
def main():
    st.title("🚀 問卷快速輸入系統")
    
    # 自動建立檔案處理：如果不存在就建新的，確保免上傳模板
    if not os.path.exists(OUTPUT_PATH):
        create_new_workbook()
        # 清除舊的進度記錄
        if os.path.exists(PROGRESS_FILE):
            os.remove(PROGRESS_FILE)

    # 初始化 Session State
    if 'current_row' not in st.session_state:
        st.session_state.current_row = load_progress()
    if 'session_count' not in st.session_state:
        st.session_state.session_count = 0

    wb = openpyxl.load_workbook(OUTPUT_PATH)
    ws = wb.active
    
    # ── SIDEBAR: 下載與重置 ──
    with st.sidebar:
        st.header("📥 資料下載與管理")
        st.markdown("輸入完畢後，請點擊下方按鈕將結果下載回您的電腦。")
        
        # 下載按鈕
        with open(OUTPUT_PATH, "rb") as f:
            excel_data = f.read()
            
        st.download_button(
            label="📥 下載完成的問卷 Excel",
            data=excel_data,
            file_name="questionnaire_filled.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary"
        )
        
        st.divider()
        st.markdown("⚠️ **重新開始新的一批**")
        if st.button("🗑️ 刪除全部資料並重新開始", use_container_width=True):
            if os.path.exists(OUTPUT_PATH): os.remove(OUTPUT_PATH)
            if os.path.exists(PROGRESS_FILE): os.remove(PROGRESS_FILE)
            st.session_state.current_row = DATA_START_ROW
            st.session_state.session_count = 0
            st.rerun()

    seq = st.session_state.current_row - DATA_START_ROW + 1
    
    # ── UI 佈局 ──
    col1, col2 = st.columns([1, 1])
    
    with col1:
        st.info(f"📈 目前進度: **第 {seq} 份** (Excel 第 {st.session_state.current_row} 列) | 本次輸入: {st.session_state.session_count} 份")
        
        # 建立輸入表單 (clear_on_submit 讓我們送出後自動清空欄位)
        with st.form("entry_form", clear_on_submit=True):
            st.subheader("📝 資料輸入")
            choices_input = st.text_input(
                f"🔢 選擇題 ({QUESTIONS_COUNT}位數, 1~5)", 
                max_chars=QUESTIONS_COUNT,
                placeholder="例如: 1121"
            ).strip().lower()
            
            t1 = st.text_input("💬 反映與建議", placeholder="跳過請留空")
            t2 = st.text_input("💬 其他建議", placeholder="跳過請留空")
            
            st.markdown("*提示：輸入完畢後按 `Enter` 即可快速送出*")
            submitted = st.form_submit_button("🚀 送出 (Submit)", use_container_width=True)

        if submitted:
            # Validation
            if not choices_input:
                st.error("⚠️ 選擇題不能為空！")
            elif len(choices_input) != QUESTIONS_COUNT or not all(c in OPTIONS_RANGE for c in choices_input):
                st.error(f"⚠️ 格式錯誤！請輸入 {QUESTIONS_COUNT} 個 1~5 的數字 (例如: 1121)")
            else:
                choices =[int(c) for c in choices_input]
                
                # Write Choices
                for q_idx, val in enumerate(choices):
                    start_col = QUESTION_COL_START[q_idx] + 1
                    for offset in range(5):
                        ws.cell(row=st.session_state.current_row, column=start_col + offset).value = (
                            1 if (offset + 1) == val else None
                        )
                
                # Write Text
                ws.cell(row=st.session_state.current_row, column=TEXT_COL_1 + 1).value = t1 or None
                ws.cell(row=st.session_state.current_row, column=TEXT_COL_2 + 1).value = t2 or None
                
                # Save & Progress
                st.session_state.current_row += 1
                st.session_state.session_count += 1
                
                wb.save(OUTPUT_PATH)
                save_progress(st.session_state.current_row)
                
                st.success(f"✅ 成功儲存第 {seq} 份！")
                st.rerun() # 重新刷新畫面以更新進度

    with col2:
        st.subheader("📊 即時統計報表")
        stats_df = update_and_get_stats(wb, st.session_state.current_row)
        
        if stats_df is not None:
            # 隱藏 pandas 的預設 index 呈現
            st.dataframe(stats_df, hide_index=True, use_container_width=True)
        else:
            st.write("⚠️ 目前尚無資料可統計。開始輸入以查看圖表！")

if __name__ == "__main__":
    main()